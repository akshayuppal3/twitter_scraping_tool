####################
##Class containing##
# helper functions###
####################
from time import sleep
import tweepy
import os
import argparse
import pandas as pd
import pandas.io.common
from pathlib import Path
import json
from openpyxl import load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError
import posixpath
import numpy as np
import nltk
import ast
from setup import setup_env
from tqdm import tqdm
tqdm.pandas()

## loading the config file
dir_name = os.getcwd()
path1 = str(Path(os.getcwd()).parent.parent)
filepath = posixpath.join(path1, 'config.json')
with open(filepath) as f:
	data = json.load(f)
logdir = os.path.join(path1, data['logdir'])
twintDir = os.path.join(path1, data['twintdir'])
inputdir = os.path.join(path1, data['inputdir'])
modeldir = os.path.join(path1, data['modeldir'])

## logging informaation
format = "%(asctime)s %(levelname)-8s %(message)s"
dateFormat = "%Y-%m-%d"
testLimit = 5
twintLimit = 1
userTimelineLimit = 200  # limit for the no of tweets extracted from user timeline
friendLimit = 100
startDate = '2018-05-01'
endDate = '2018-05-02'

setup_env()  # download necessary nltk packages
stopwords = nltk.corpus.stopwords.words('english')


# @param dataframe and output filename
def output_to_csv(df, filename):
	if (df is not None and not df.empty):
		df.to_csv(filename, sep=",", line_terminator='\n', index=None)
	else:
		print("datframe is empty")


# conversion of str to bool
def str2bool(v):
	if v.lower() in ('train', 't', 't', 'y', '1'):
		return True
	elif v.lower() in ('no', 'false', 'f', 'n', '0'):
		return False
	else:
		raise argparse.ArgumentTypeError('Boolean value expected.')

# @Deprecated
# handle the rate limit (wait for 15min (API constarints))
def limit_handler(cursor):
	while True:
		try:
			yield cursor.next()
		except tweepy.RateLimitError:
			print("sleeping -rate limit exceeded")
			sleep(15)


# @param df
# @return list of unique users
def getUsers(df, type):
	if (df is not None and not df.empty):
		if (type == 'ID'):
			if ('userID' in df):
				unique_users = list(df['userID'].unique())
				return unique_users
			else:
				return None
		elif (type == 'name'):
			if ('userName' in df):
				unique_names = list(df['userName'].unique())
				return unique_names
			else:
				return None
	else:
		return None


# read CSV to generate dataframe
# @return df
def readCSV(path):
	try:
		df = pd.read_csv(path, lineterminator='\n', index_col=None)
		if "userName\r" in df:  # windows problem
			df["userName\r"] = df["userName\r"].str.replace(r'\r', '')
			df.rename(columns={'userName\r': "userName"}, inplace=True)
		return df
	except FileNotFoundError:
		print("[ERROR] file not found")
	except pd.io.common.EmptyDataError:
		print("[ERROR] empty file")


# read xlsx file as csv and return dataframe
# @return df @param filepath
def read_excel(path):
	try:
		df = pd.read_excel(path, 'Sheet1', index_col=None)
		return df
	except FileNotFoundError:
		print("[ERROR] file not found")
	except pd.io.common.EmptyDataError:
		print("[ERROR] empty file")


# Convert df to excel
# appends to the excel file path specified(or create a nex file with that name)
def df_write_excel(df, filepath):
	# df = df.applymap(lambda x: x.encode('unicode_escape').
	#                  decode('utf-8') if isinstance(x, str) else x)             # prevent Illegal character errror
	try:
		writer = pd.ExcelWriter(filepath, engine='openpyxl')
		if os.path.isfile(filepath):
			writer.book = load_workbook(filepath)
			writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
			max_row = writer.book.active.max_row
			sheetname = writer.book.active.title
			if len(df > 1):
				for index, row in df.iterrows():
					try:
						row.to_excel(writer, sheet_name=sheetname, startrow=max_row, index=True, header=False)
					except IllegalCharacterError:
						print("Illegal character error")
						continue
			else:
				df.to_excel(writer, sheet_name=sheetname, startrow=max_row, index=False, header=False)
		else:
			df.to_excel(writer, index=False)  # in case the file does not exists
		try:
			writer.save()
		except OSError:
			print("File is open: or permission denied")
	except IllegalCharacterError:
		print("Illegal character error")


# appends to existing csv file
def df_write_csv(df, filepath):
	try:
		if os.path.isfile(filepath):
			with open(filepath, 'a') as f:
				df.to_csv(f,header=False)
		else:
			df.to_csv(filepath)  # in case the file does not exists
	except IllegalCharacterError:
		print("Illegal character error")

## @param df  , @return length of hashtags
def hashtag_count(df):
	if df is not np.nan:
		hashtags = ast.literal_eval(df)
		if (hashtags is not None):
			return (len(hashtags))
		else:
			return (0)
	else:
		return (0)

## return tokenize sentences
# @param df, columns
# @returns sentences
def get_sentences(df, column):
	sentences = list(df[column].apply(get_tokenize_words))
	return (sentences)

# function to get the nearest postion for element in the list
def nearest(items, pivot):
	return min(items, key=lambda x: abs(x - pivot))


# function to return tokenize words
## @param text @return tokenize words
def get_tokenize_words(text):
	tkz = nltk.tokenize
	words = tkz.word_tokenize(text)
	words = [ele for ele in words if ((ele not in stopwords) and len(ele) > 2 and (ele.isalpha()))]
	words = [word.lower() for word in words]
	return words


def getHashtags(tweetObj, extended=False):
	if extended == True:
		if 'retweeted_status' in tweetObj._json.keys():
			if len(tweetObj.retweeted_status.entities['hashtags']) != 0:
				hashtags = [i['text'] for i in tweetObj.retweeted_status.entities['hashtags']]
			else:
				hashtags = None
		else:
			hashtags = getHashtags(tweetObj,extended=False)
	else:
		if len(tweetObj.entities['hashtags']) != 0:
			hashtags = [j['text'] for j in tweetObj.entities['hashtags']]
		else:
			hashtags = "None"
	return hashtags

# @params passing tweet, friendOpt and userinfo(in case of following)
# returns data frame of tweet and user info
def getTweetObject(tweetObj, parentID=None):
	if parentID is not None:
		data = pd.DataFrame.from_records(
			[{
				'userID': tweetObj.id,
				'parentID': parentID,
				'userName': tweetObj.name,
				'userDescription': tweetObj.description,
				'userCreatedAt': tweetObj.created_at,
				'userLocation': tweetObj.location,
				'favourites_count': tweetObj.favourites_count,
				'friendsCount': tweetObj.friends_count,
				'userFollowersCount': tweetObj.followers_count,
				'listedCount': tweetObj.listed_count,
				'lang': tweetObj.lang,
				'url': tweetObj.url,
				'imageurl': tweetObj.profile_image_url,
				'userVerified': tweetObj.verified,
				'isProtected': tweetObj.protected,
				'notifications': tweetObj.notifications,
				'statusesCount': tweetObj.statuses_count,
				'geoEnabled': tweetObj.geo_enabled,
				'contributorEnabled': tweetObj.contributors_enabled,
				# 'status': tweetObj.status,
				'withheldinCountries': tweetObj.withheld_in_countries if 'withheld_in_countries' in tweetObj._json.keys() else None,
				'withheldScope': tweetObj.withheld_scope if 'withheld_scope' in tweetObj._json.keys() else None,
			}], index=[0])
	else:
		if 'retweeted_status' in tweetObj._json.keys():
			retweetText = tweetObj.retweeted_status.full_text.replace("\n", " ")
			text = tweetObj.full_text.replace("\n", " ")
			hashtags = getHashtags(tweetObj, extended=True)  # for retweeted status
			retweeted = True
		else:
			text = tweetObj.full_text.replace("\n", " ")
			retweetText = None
			hashtags = getHashtags(tweetObj)
			retweeted = False
		data = pd.DataFrame.from_records(
			[{
				'tweetId': tweetObj.id_str,
				'tweetText': text,
				'retweetText': retweetText,
				'retweetCount': tweetObj.retweet_count,
				'retweeted': retweeted,
				'tweetCreatedAt': tweetObj.created_at,
				'userName': tweetObj.user.name,
				'userID': tweetObj.user.id,
				'userCreatedAt': tweetObj.user.created_at,
				'userLocation': tweetObj.user.location,
				'userDescription': tweetObj.user.description.replace("\n", " "),
				'friendsCount': tweetObj.user.friends_count,
				'followersCount': tweetObj.user.followers_count,
				'listedCount': tweetObj.user.listed_count,
				'lang': tweetObj.lang,
				'hashtags': hashtags,
				'url': tweetObj.user.url,
				'imageurl': tweetObj.user.profile_image_url,
			}], index=[0])
	return (data)
